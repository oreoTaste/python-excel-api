# pip install openpyxl requests python-dotenv
# pyinstaller --onefile --noconsole main.py
import os
import json
import openpyxl
import requests
import tkinter as tk
from tkinter import messagebox, filedialog, scrolledtext
from tkinter import ttk  # 현대적인 테마 위젯을 위해 사용
from dotenv import load_dotenv

# --- [1. 설정 및 환경 변수 관리] ---
load_dotenv()
CONFIG_FILE = "settings.json"

# 기본 설정값
DEFAULT_CONFIG = {
    "file_path": "contract_data.xlsx",
    "src_sheet": "작업리스트",
    "tgt_sheet": "완료리스트",
    "sum_formula_cell": "B14",
    "user_id": os.getenv("COMPANY_ID", "your_id"),
    "api_login_url": os.getenv("LOGIN_URL", "https://company.com/api/login"),
    "api_save_url": os.getenv("SAVE_URL", "https://company.com/api/save")
}

def load_settings():
    """설정 파일(settings.json)을 로드합니다."""
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except Exception:
                return DEFAULT_CONFIG
    return DEFAULT_CONFIG

def save_settings(config):
    """현재 설정을 파일로 저장합니다."""
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

# --- [2. 메인 애플리케이션 클래스] ---

class AutomationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("업무 자동화 도구 v2.4")
        self.root.geometry("550x750")
        
        # 로드된 설정 적용
        self.config = load_settings()
        self.entries = {}
        
        # UI 스타일 설정 (ttk)
        self.style = ttk.Style()
        self.style.configure("TLabel", font=("Malgun Gothic", 10))
        self.style.configure("TButton", font=("Malgun Gothic", 10))
        
        self.setup_ui()
        
        # [개선 1] 프로그램 실행 시 업체명 입력창에 커서(포커스) 자동 배치
        if "name" in self.entries:
            self.entries["name"].focus_set()

        # [개선 2] 어디서든 엔터 키를 누르면 실행되도록 전역 바인딩
        # lambda event를 사용하여 이벤트 인자를 무시하고 실행 함수를 호출합니다.
        self.root.bind('<Return>', lambda event: self.start_process())

    def setup_ui(self):
        """전체 UI 레이아웃을 구성합니다."""
        main_container = ttk.Frame(self.root, padding="20")
        main_container.pack(fill="both", expand=True)

        # 1. 상단: 파일 및 시트 설정 영역
        self._create_config_ui(main_container)
        
        # 2. 중간: 데이터 입력 영역
        self._create_input_ui(main_container)

        # 3. 버튼: 실행 버튼
        self.btn_run = tk.Button(
            main_container, text="실행 및 설정 저장 (Enter)", command=self.start_process, 
            bg="#27ae60", fg="white", font=("Malgun Gothic", 12, "bold"), 
            height=2, relief="flat", cursor="hand2"
        )
        self.btn_run.pack(pady=15, fill="x")

        # 4. 하단: 실시간 로그 콘솔
        self._create_log_ui(main_container)

    def _create_config_ui(self, parent):
        """엑셀 파일 경로 및 시트 설정 영역"""
        frame = ttk.LabelFrame(parent, text=" 기본 설정 (자동 업데이트) ", padding="10")
        frame.pack(fill="x", pady=(0, 10))

        ttk.Label(frame, text="엑셀 파일:").grid(row=0, column=0, sticky="e", pady=2)
        self.ent_file = ttk.Entry(frame, width=35)
        self.ent_file.insert(0, self.config.get("file_path", ""))
        self.ent_file.grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="찾기", width=7, command=self.browse_file).grid(row=0, column=2)

        self.ent_src = self._add_config_row(frame, "소스 시트:", self.config.get("src_sheet"), 1)
        self.ent_tgt = self._add_config_row(frame, "타겟 시트:", self.config.get("tgt_sheet"), 2)
        self.ent_sum_cell = self._add_config_row(frame, "합계 수식 셀:", self.config.get("sum_formula_cell"), 3)
        
        ttk.Label(frame, text="* 작업 성공 시 수식 위치가 자동으로 갱신됩니다.", 
                  foreground="#2980b9", font=("Malgun Gothic", 8)).grid(row=4, column=1, sticky="w")

    def _add_config_row(self, frame, label, value, row):
        ttk.Label(frame, text=label).grid(row=row, column=0, sticky="e", pady=5)
        entry = ttk.Entry(frame, width=43)
        entry.insert(0, value)
        entry.grid(row=row, column=1, columnspan=2, padx=5, sticky="w")
        return entry

    def _create_input_ui(self, parent):
        """데이터 입력 필드 영역"""
        frame = ttk.LabelFrame(parent, text=" 작업 데이터 입력 ", padding="10")
        frame.pack(fill="x", pady=5)

        fields = [
            ("업체명(검색)", "name", "(대소문자 무관 검색)"),
            ("계약금액", "amount", "(숫자만 입력)"),
            ("계약날짜", "cdate", "(YYYY-MM-DD)"),
            ("시작날짜", "sdate", "(YYYY-MM-DD)")
        ]

        for i, (label, key, guide) in enumerate(fields):
            ttk.Label(frame, text=f"{label}:").grid(row=i, column=0, sticky="e", pady=5)
            ent = ttk.Entry(frame, width=28)
            ent.grid(row=i, column=1, padx=5, sticky="w")
            self.entries[key] = ent
            ttk.Label(frame, text=guide, foreground="#7f8c8d", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky="w")

    def _create_log_ui(self, parent):
        frame = ttk.LabelFrame(parent, text=" 처리 로그 ", padding="5")
        frame.pack(fill="both", expand=True)
        
        self.log_area = scrolledtext.ScrolledText(frame, height=10, font=("Consolas", 9), state='disabled', bg="#f8f9fa")
        self.log_area.pack(fill="both", expand=True)
        self.write_log("시스템이 준비되었습니다. 엔터 키로 실행 가능합니다.")

    def write_log(self, message):
        self.log_area.config(state='normal')
        self.log_area.insert(tk.END, f"[{self._get_now()}] {message}\n")
        self.log_area.see(tk.END)
        self.log_area.config(state='disabled')

    def _get_now(self):
        import datetime
        return datetime.datetime.now().strftime("%H:%M:%S")

    # --- [3. 비즈니스 로직] ---

    def browse_file(self):
        file_selected = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_selected:
            self.ent_file.delete(0, tk.END)
            self.ent_file.insert(0, file_selected)
            self.write_log(f"파일 선택: {os.path.basename(file_selected)}")

    def save_current_settings(self):
        self.config.update({
            "file_path": self.ent_file.get(),
            "src_sheet": self.ent_src.get(),
            "tgt_sheet": self.ent_tgt.get(),
            "sum_formula_cell": self.ent_sum_cell.get()
        })
        save_settings(self.config)

    def start_process(self):
        """작업 시작: 검색 전 사용자에게 일차 확인을 받습니다."""
        
        # [개선 3] 실행 전 확인 절차 (엔터 실수 방지용)
        if not messagebox.askyesno("실행 확인", "설정을 저장하고 데이터를 검색하시겠습니까?"):
            self.write_log("사용자가 실행을 취소했습니다.")
            return

        self.save_current_settings()
        keyword = self.entries["name"].get().strip()
        
        if not keyword:
            messagebox.showwarning("알림", "검색할 업체명을 입력하세요.")
            self.entries["name"].focus_set()
            return

        self.write_log(f"'{keyword}' 검색을 시작합니다...")
        try:
            wb = openpyxl.load_workbook(self.config["file_path"])
            ws = wb[self.config["src_sheet"]]
            matches = self._search_keyword_in_sheet(ws, keyword)
            
            if not matches:
                self.write_log(f"검색 실패: '{keyword}' 결과 없음")
                messagebox.showinfo("결과", f"'{keyword}'와 일치하는 데이터를 찾지 못했습니다.")
            elif len(matches) == 1:
                self.confirm_and_run(matches[0])
            else:
                self.write_log(f"다중 결과 발생: {len(matches)}건")
                self.show_selection_window(matches)
        except Exception as e:
            self.write_log(f"오류 발생: {str(e)}")
            messagebox.showerror("오류", f"처리 중 에러 발생:\n{str(e)}")

    def _search_keyword_in_sheet(self, sheet, keyword):
        """대소문자 구분 없이 검색 수행"""
        results = []
        search_key = keyword.lower()
        for row in sheet.iter_rows(min_row=2):
            val = str(row[0].value).lower() if row[0].value else ""
            if search_key in val:
                results.append({"idx": row[0].row, "data": [c.value for c in row]})
        return results

    def show_selection_window(self, matches):
        """여러 업체가 검색된 경우 선택 창 표시"""
        win = tk.Toplevel(self.root)
        win.title("업체 선택")
        win.geometry("300x400")
        
        ttk.Label(win, text="처리할 대상을 선택하세요:", padding=10).pack()
        lb = tk.Listbox(win, font=("Malgun Gothic", 10))
        lb.pack(padx=10, pady=5, fill="both", expand=True)
        
        for m in matches:
            lb.insert(tk.END, f"{m['data'][0]} ({m['data'][1]}원)")
            
        def on_select():
            if lb.curselection():
                idx = lb.curselection()[0]
                win.destroy()
                self.confirm_and_run(matches[idx])
        
        ttk.Button(win, text="선택 완료", command=on_select).pack(pady=10)

    def confirm_and_run(self, match):
        """최종 데이터 이동 및 API 전송 확인"""
        # 개별 데이터에 대한 최종 확인 (이중 안전 장치)
        if not messagebox.askyesno("최종 확인", f"선택된 업체: '{match['data'][0]}'\n\n이 데이터를 이동하고 서버로 전송할까요?"):
            self.write_log("데이터 이동 취소됨.")
            return

        try:
            self.write_log("엑셀 데이터 이동 작업 중...")
            new_formula_addr = self._execute_excel_update(match)
            
            self.ent_sum_cell.delete(0, tk.END)
            self.ent_sum_cell.insert(0, new_formula_addr)
            self.save_current_settings()
            
            self.write_log("API 서버 전송 중...")
            success, msg = self.call_api_service(match['data'])
            
            if success:
                self.write_log("성공: 데이터 이동 및 서버 전송이 완료되었습니다.")
                messagebox.showinfo("성공", "모든 작업이 완료되었습니다!")
                # 다음 입력을 위해 입력란 초기화 및 포커스
                self.entries["name"].delete(0, tk.END)
                self.entries["name"].focus_set()
            else:
                self.write_log(f"주의: {msg}")
                messagebox.showwarning("부분 성공", f"엑셀은 수정되었으나 서버 전송 실패:\n{msg}")

        except Exception as e:
            self.write_log(f"실행 오류: {str(e)}")
            messagebox.showerror("실행 오류", f"작업 중 중단됨: {str(e)}")

    def _execute_excel_update(self, match):
        """엑셀 시트 간 데이터 이동 및 수식 갱신"""
        wb = openpyxl.load_workbook(self.config["file_path"])
        src_ws = wb[self.config["src_sheet"]]
        tgt_ws = wb[self.config["tgt_sheet"]]
        
        addr = self.config["sum_formula_cell"].strip()
        target_cell = tgt_ws[addr]
        sum_row, sum_col = target_cell.row, target_cell.column
        
        insert_idx = sum_row
        src_ws.delete_rows(match['idx'])
        tgt_ws.insert_rows(insert_idx)
        
        for c_idx, val in enumerate(match['data'], 1):
            tgt_ws.cell(row=insert_idx, column=c_idx, value=val)

        new_sum_row = sum_row + 1
        col_letter = openpyxl.utils.get_column_letter(sum_col)
        new_formula = f"=SUM({col_letter}2:{col_letter}{new_sum_row - 1})"
        tgt_ws.cell(row=new_sum_row, column=sum_col, value=new_formula)
        
        wb.save(self.config["file_path"])
        return f"{col_letter}{new_sum_row}"

    def call_api_service(self, data):
        """API 서버 데이터 전송"""
        try:
            payload = {
                "companyName": data[0], "contractAmount": data[1],
                "contractDate": data[2], "startDate": data[3]
            }
            with requests.Session() as s:
                s.post(self.config["api_login_url"], 
                       data={"userId": self.config["user_id"], "password": os.getenv("COMPANY_PW", "default_pw")},
                       timeout=10).raise_for_status()
                s.post(self.config["api_save_url"], json=payload, timeout=10).raise_for_status()
                return True, "API 성공"
        except Exception as e:
            return False, str(e)

if __name__ == "__main__":
    root = tk.Tk()
    app = AutomationApp(root)
    root.mainloop()